"""Excel 文档加载器"""
from pathlib import Path
from typing import List, Union, Optional


def read_excel(file_path: Union[str, Path], max_rows: Optional[int] = None) -> str:
    """
    读取 Excel 文件内容，返回纯文本格式
    
    Args:
        file_path: Excel 文件路径
        max_rows: 最大读取行数（用于控制上下文长度）
    
    Returns:
        Excel 内容的纯文本表示
    """
    try:
        import openpyxl
    except ImportError:
        print("[WARN] openpyxl 未安装，无法读取 Excel 文件")
        return ""
    
    file_path = Path(file_path)
    if not file_path.exists():
        return ""
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        content_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content_parts.append(f"【工作表: {sheet_name}】")
            
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                # 过滤空行
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):
                    content_parts.append(" | ".join(row_values))
                    row_count += 1
                    
                    if max_rows and row_count >= max_rows:
                        content_parts.append(f"...（已截断，共 {sheet.max_row} 行）")
                        break
        
        wb.close()
        return "\n".join(content_parts)
        
    except Exception as e:
        print(f"[ERROR] Excel 读取失败 {file_path}: {e}")
        return ""


def load_excel_document(file_path: Union[str, Path], max_rows: Optional[int] = None) -> dict:
    """
    加载 Excel 文件，返回结构化数据
    
    Returns:
        {
            "content": str,  # 纯文本内容
            "metadata": {
                "source": str,      # 文件路径
                "title": str,       # 文件名（不含扩展名）
                "sheets": list,     # 工作表列表
                "row_count": int    # 总行数
            }
        }
    """
    file_path = Path(file_path)
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        content = read_excel(file_path, max_rows)
        sheet_names = wb.sheetnames
        total_rows = sum(sheet.max_row for sheet in [wb[name] for name in sheet_names])
        
        wb.close()
        
        return {
            "content": content,
            "metadata": {
                "source": str(file_path),
                "title": file_path.stem,
                "sheets": sheet_names,
                "row_count": total_rows
            }
        }
    except Exception as e:
        return {
            "content": "",
            "metadata": {
                "source": str(file_path),
                "title": file_path.stem,
                "sheets": [],
                "row_count": 0,
                "error": str(e)
            }
        }


def load_excel_directory(directory: Union[str, Path], max_rows_per_file: Optional[int] = None) -> List[dict]:
    """
    加载目录下所有 Excel 文件
    
    Args:
        directory: 目录路径
        max_rows_per_file: 每个文件最大读取行数
    
    Returns:
        Excel 文档列表
    """
    directory = Path(directory)
    documents = []
    
    if not directory.exists():
        return documents
    
    for excel_file in directory.rglob("*.xlsx"):
        try:
            doc = load_excel_document(excel_file, max_rows_per_file)
            if doc["content"]:
                documents.append(doc)
                print(f"[OK] 加载 Excel: {excel_file.name}")
        except Exception as e:
            print(f"[FAIL] Excel 加载失败 {excel_file.name}: {e}")
    
    for excel_file in directory.rglob("*.xls"):
        try:
            doc = load_excel_document(excel_file, max_rows_per_file)
            if doc["content"]:
                documents.append(doc)
                print(f"[OK] 加载 Excel: {excel_file.name}")
        except Exception as e:
            print(f"[FAIL] Excel 加载失败 {excel_file.name}: {e}")
    
    return documents
